from ProcessPhasesNps import process_phases_nps
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def update_excel_e_nps(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Lista de títulos para os cabeçalhos conforme especificado
    headers = [
        "Período:", "NPS", "Grupo minoritário?", "Cargo:",
        "Sinto-me envolvido com o trabalho que faço.",
        "Estou entusiasmado com meu trabalho.",
        "Em meu trabalho, sinto-me cheio de energia.",
        "Eu entendo como meu trabalho contribui para o alcance das metas e objetivos da empresa.",
        "Eu sinto que faço a diferença no meu time.",
        "Eu sinto que, se eu cometer um erro, isso não se voltará contra mim.",
        "Sinto que a cultura da UCJ está alinhada com as minhas crenças e valores.",
        "A UCJ possui lideranças com as quais me identifico.",
        "Sinto que a minha liderança direta se preocupa comigo como pessoa.",
        "Sinto que a minha liderança direta constrói um ambiente positivo, ou seja, temos uma comunicação aberta e transparente, falamos de dificuldades e temos uma cultura de feedbacks constantes.",
        "Eu estou satisfeito em relação ao tempo que dedico para o meu trabalho, meus estudos, minha família, meus amigos e minha saúde.",
        "Eu sinto que a minha liderança direta encoraja e apoia meu desenvolvimento.",
        "Sinto que tenho voz ativa opinar e fazer acontecer as transformações em que acredito.",
        "Sinto que sou comunicado (a) das informações relevantes para o meu trabalho e sobre assuntos gerais relevantes na empresa.",
        "Sinto que o ambiente em que trabalho colabora para a minha produtividade.",
        "O que motivou sua resposta.",
        "NPS produtos", "Comente sobre o que motivou essa resposta.",
        "O que podemos fazer para melhorar enquanto empresa?", "Qual(is)?"
    ]


    # Aplicando os cabeçalhos e seus estilos
    def setup_headers(ws, headers):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


    # Ajuste das colunas conforme anteriormente
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in col), default=0)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
    
    standard_width = 8.43
    for col in range(5, 20):  # Colunas E(5) até S(19)
        ws.column_dimensions[get_column_letter(col)].width = standard_width
    
    column_letter = get_column_letter(2)  # 'B' corresponde ao índice 2
    desired_width = 9  # Você pode ajustar esse valor conforme necessário

    ws.column_dimensions[column_letter].width = desired_width

    column_t_index = headers.index("O que motivou sua resposta.") + 1  
    t_header_value = ws.cell(row=1, column=column_t_index).value
    ws.column_dimensions[get_column_letter(column_t_index)].width = len(t_header_value) + 2

    column_t_index = headers.index("Comente sobre o que motivou essa resposta.") + 1  
    t_header_value = ws.cell(row=1, column=column_t_index).value
    ws.column_dimensions[get_column_letter(column_t_index)].width = len(t_header_value) + 2
    
    column_t_index = headers.index("O que podemos fazer para melhorar enquanto empresa?") + 1  
    t_header_value = ws.cell(row=1, column=column_t_index).value
    ws.column_dimensions[get_column_letter(column_t_index)].width = len(t_header_value) + 2

    # Configurando a primeira linha com cabeçalhos
    setup_headers(ws, headers)

    # Processa as fases e atualiza a planilha
    row_num = 2
    ws, row_num = process_phases_nps(all_phases, headers, ws, row_num)


    print(f"Dados atualizados na aba '{sheet_name}'.")
