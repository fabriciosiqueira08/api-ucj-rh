from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def process_phases_nps(all_phases, headers, ws, row_num):
    # Estilo de fonte para o restante das células
    normal_font = Font(name='Arial', size=11, bold=False)
    alignment_bottom = Alignment(vertical='bottom')

    index_col_t = headers.index("O que motivou sua resposta.") + 1
    index_col_v = headers.index("Comente sobre o que motivou essa resposta.") + 1
    index_col_w = headers.index("O que podemos fazer para melhorar enquanto empresa?") + 1

    columns_with_decimals = ['C'] + [chr(i) for i in range(ord('F'), ord('T') + 1)]

    for phase in all_phases:
        num_trimestre = ''
        year = ''

        if "Tri" in phase['name']:
            num_trimestre = phase['name'].split()[0]  # Assume que o formato é "X° Tri YYYY"
            year = phase['name'].split()[-1]
            try:
                year = int(year)
            except ValueError:
                year = None
            formatted_name = f"{num_trimestre} Trimestre"
        else:
            formatted_name = phase['name']

        if phase['name'] not in ["Caixa de entrada", "Concluído"]:
            for card_edge in phase['cards']['edges']:
                card = card_edge['node']
                field_values = {header: "" for header in headers}  # Inicializa todos os campos com string vazia
                field_values["Período:"] = formatted_name
                field_values["ANO:"] = year

                try:
                    field_values["NPS"] = int(card['title'])
                except ValueError:
                    try:
                        field_values["NPS"] = float(card['title'])
                    except ValueError:
                        field_values["NPS"] = card['title'] 

                for field in card['fields']:
                    header_name = next((h for h in headers if h.endswith(field['name'] + ":")), None)
                    if field['name'] in headers:
                        # Convertendo para número se possível, senão mantém como string
                        try: 
                            field_values[field['name']] = int(field['value']) if field['value'].isdigit() else float(field['value'])
                        except ValueError:
                            field_values[field['name']] = field['value']  # Preenche os valores onde o cabeçalho coincide com o nome do campo
                    elif header_name:
                        try:
                            field_values[header_name] = int(field['value']) if field['value'].isdigit() else float(field['value'])
                        except ValueError:
                            field_values[header_name] = field['value']  
                    elif field['name'] == "Você pertence a algum grupo minoritário?":
                        field_values["Grupo minoritário?"] = field['value']
                    elif field['name'] == "Você é:":
                        field_values["Cargo:"] = field['value']
                    elif field['name'] == "Em uma escala de 0 a 10, o quanto você recomendaria os produtos da UCJ para amigos ou familiares?":
                        try:
                             # Certificando-se de que o valor é um inteiro
                            field_values["NPS produtos"] = int(field['value'])
                        except ValueError:
                            field_values["NPS produtos"] = field['value']
                    elif field['name'] == "Comente sobre o que motivou sua resposta.":
                        field_values["O que motivou sua resposta."] = field['value']
                # Preenche a linha com os valores coletados
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=field_values[header])
                    cell.font = normal_font
                    cell.alignment = alignment_bottom

                    if get_column_letter(col_num) in columns_with_decimals:
                        cell.number_format = '0.00'

                    if header == "ANO:" and isinstance(field_values[header], int):
                        cell.number_format = '0'  # Define formatação como inteiro sem casas decimais

                    if col_num in [index_col_t, index_col_v, index_col_w]:
                        cell.alignment = Alignment(wrap_text=True)  # Aplica quebra de texto nas colunas específicas


                row_num += 1 

    return ws, row_num
