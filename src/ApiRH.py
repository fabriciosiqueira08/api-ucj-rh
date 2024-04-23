import requests, os
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Definições
load_dotenv()
PIPEFY_API_TOKEN = os.getenv('PIPEFY_API_TOKEN')
PIPEFY_GRAPHQL_ENDPOINT = os.getenv('PIPEFY_GRAPHQL_ENDPOINT')
PIPE_TO_FILE = {
    'RH - E-NPS': ('[DRH] ENPS.xlsx', 'Banco de dados'),
    'RH - Matriz de Cursos': ('[DRH] Matriz de Cursos.xlsx', 'Banco de dados'),
    'RH - Painel Controle Membros': ('[UCJ] Relação de Membros - 2024.xlsx', 'Banco de dados')
}

# IDs dos Pipes 
PIPE_IDS = {
    'RH - E-NPS': '301823995',
    'RH - Matriz de Cursos': '301682389',
    'RH - Painel Controle Membros': '301654957'
}

# Função para consultar os dados do Pipefy
def fetch_pipefy_data(pipe_id, cursor=None, page_size=30,):
    # Cursor clause to handle pagination
    cursor_clause = f', after: "{cursor}"' if cursor else ""
    
    query = f"""
    query {{
      pipe(id: "{pipe_id}") {{
        phases {{
          name
          cards(first: {page_size}{cursor_clause}) {{
            edges {{
              node {{
                title
                fields {{
                  name
                  value
                }}
              }}
            }}
            pageInfo {{
              hasNextPage
              endCursor
            }}
          }}
        }}
      }}
    }}
    """

    headers = {'Authorization': f'Bearer {PIPEFY_API_TOKEN}'}
    response = requests.post(PIPEFY_GRAPHQL_ENDPOINT, json={'query': query}, headers=headers)
    data = response.json()
    return data


def fetch_all_cards(pipe_id):
    all_phases = []
    initial_data = fetch_pipefy_data(pipe_id)  # Esta função precisa tratar corretamente o cursor inicial como None

    for phase_data in initial_data['data']['pipe']['phases']:

        if phase_data['name'] == "Histórico":
            continue  # Pula a fase "Histórico"

        phase = phase_data
        cards = phase_data['cards']['edges']
        pageInfo = phase_data['cards']['pageInfo']
        cursor = pageInfo['endCursor']

        # Assegurar que a fase atual seja corretamente paginada
        while pageInfo['hasNextPage']:
            more_data = fetch_pipefy_data(pipe_id, cursor)  # Adicionar parâmetro para fase
            more_phase_data = next((p for p in more_data['data']['pipe']['phases'] if p['name'] == phase['name']), None)
            if more_phase_data:
                more_cards = more_phase_data['cards']['edges']
                cards.extend(more_cards)
                pageInfo = more_phase_data['cards']['pageInfo']
                cursor = pageInfo['endCursor']
            else:
                break  # Interrompe se a fase específica não for encontrada

        phase['cards']['edges'] = cards
        all_phases.append(phase)
        print(f"Fase '{phase['name']}' processada com {len(cards)} cartões.")

    return all_phases

# Função para limpar os valores de campo de lista
def clean_value(value):
    if isinstance(value, str) and value.startswith("[\"") and value.endswith("\"]"):
        # Assume que há apenas um item na lista e remove os caracteres indesejados
        return value[2:-2]  # Remove os dois primeiros e os dois últimos caracteres
    return value


def process_phases_nps(all_phases, headers, ws, row_num):
    # Estilo de fonte para o restante das células
    normal_font = Font(name='Arial', size=11, bold=False)
    alignment_bottom = Alignment(vertical='bottom')

    index_col_t = headers.index("O que motivou sua resposta.") + 1
    index_col_v = headers.index("Comente sobre o que motivou essa resposta.") + 1
    index_col_w = headers.index("O que podemos fazer para melhorar enquanto empresa?") + 1


    for phase in all_phases:
        if "Tri" in phase['name']:
            num_trimestre = phase['name'].split()[0]  # Assume que o formato é "X° Tri YYYY"
            formatted_name = f"{num_trimestre} Trimestre"
        else:
            formatted_name = phase['name']
        if phase['name'] not in ["Caixa de entrada", "Concluído"]:
            for card_edge in phase['cards']['edges']:
                card = card_edge['node']
                field_values = {header: "" for header in headers}  # Inicializa todos os campos com string vazia
                field_values["Período:"] = formatted_name

                try:
                    field_values["NPS"] = int(card['title'])
                except ValueError:
                    try:
                        field_values["NPS"] = float(card['title'])
                    except ValueError:
                        field_values["NPS"] = card['title'] 

                for field in card['fields']:
                    header_name = next((h for h in headers if h.endswith(field['name'] + ":")), None)
                    if field['name'] in headers:
                        # Convertendo para número se possível, senão mantém como string
                        try: 
                            field_values[field['name']] = int(field['value']) if field['value'].isdigit() else float(field['value'])
                        except ValueError:
                            field_values[field['name']] = field['value']  # Preenche os valores onde o cabeçalho coincide com o nome do campo
                    elif header_name:
                        try:
                            field_values[header_name] = int(field['value']) if field['value'].isdigit() else float(field['value'])
                        except ValueError:
                            field_values[header_name] = field['value']  
                    elif field['name'] == "Você pertence a algum grupo minoritário?":
                        field_values["Grupo minoritário?"] = field['value']
                    elif field['name'] == "Você é:":
                        field_values["Cargo:"] = field['value']
                    elif field['name'] == "Em uma escala de 0 a 10, o quanto você recomendaria os produtos da UCJ para amigos ou familiares?":
                        try:
                             # Certificando-se de que o valor é um inteiro
                            field_values["NPS produtos"] = int(field['value'])
                        except ValueError:
                            field_values["NPS produtos"] = field['value']
                    elif field['name'] == "Comente sobre o que motivou sua resposta.":
                        field_values["O que motivou sua resposta."] = field['value']
                # Preenche a linha com os valores coletados
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=field_values[header])
                    cell.font = normal_font
                    cell.alignment = alignment_bottom
                    if col_num in [index_col_t, index_col_v, index_col_w]:
                        cell.alignment = Alignment(wrap_text=True)  # Aplica quebra de texto nas colunas específicas


                row_num += 1 

    return ws, row_num

# Função para criar/atualizar o arquivo do Excel com os dados do Pipefy
def update_excel_e_nps(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Lista de títulos para os cabeçalhos conforme especificado
    headers = [
        "Período:", "NPS", "Grupo minoritário?", "Cargo:",
        "Sinto-me envolvido com o trabalho que faço.",
        "Estou entusiasmado com meu trabalho.",
        "Em meu trabalho, sinto-me cheio de energia.",
        "Eu entendo como meu trabalho contribui para o alcance das metas e objetivos da empresa.",
        "Eu sinto que faço a diferença no meu time.",
        "Eu sinto que, se eu cometer um erro, isso não se voltará contra mim.",
        "Sinto que a cultura da UCJ está alinhada com as minhas crenças e valores.",
        "A UCJ possui lideranças com as quais me identifico.",
        "Sinto que a minha liderança direta se preocupa comigo como pessoa.",
        "Sinto que a minha liderança direta constrói um ambiente positivo, ou seja, temos uma comunicação aberta e transparente, falamos de dificuldades e temos uma cultura de feedbacks constantes.",
        "Eu estou satisfeito em relação ao tempo que dedico para o meu trabalho, meus estudos, minha família, meus amigos e minha saúde.",
        "Eu sinto que a minha liderança direta encoraja e apoia meu desenvolvimento.",
        "Sinto que tenho voz ativa opinar e fazer acontecer as transformações em que acredito.",
        "Sinto que sou comunicado (a) das informações relevantes para o meu trabalho e sobre assuntos gerais relevantes na empresa.",
        "Sinto que o ambiente em que trabalho colabora para a minha produtividade.",
        "O que motivou sua resposta.",
        "NPS produtos", "Comente sobre o que motivou essa resposta.",
        "O que podemos fazer para melhorar enquanto empresa?", "Qual(is)?"
    ]


    # Aplicando os cabeçalhos e seus estilos
    def setup_headers(ws, headers):
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


    # Ajuste das colunas conforme anteriormente
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in col), default=0)
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width
    
    standard_width = 8.43
    for col in range(5, 20):  # Colunas E(5) até S(19)
        ws.column_dimensions[get_column_letter(col)].width = standard_width
    
    column_letter = get_column_letter(2)  # 'B' corresponde ao índice 2
    desired_width = 9  # Você pode ajustar esse valor conforme necessário

    ws.column_dimensions[column_letter].width = desired_width

    column_t_index = headers.index("O que motivou sua resposta.") + 1  
    t_header_value = ws.cell(row=1, column=column_t_index).value
    ws.column_dimensions[get_column_letter(column_t_index)].width = len(t_header_value) + 2

    column_t_index = headers.index("Comente sobre o que motivou essa resposta.") + 1  
    t_header_value = ws.cell(row=1, column=column_t_index).value
    ws.column_dimensions[get_column_letter(column_t_index)].width = len(t_header_value) + 2
    
    column_t_index = headers.index("O que podemos fazer para melhorar enquanto empresa?") + 1  
    t_header_value = ws.cell(row=1, column=column_t_index).value
    ws.column_dimensions[get_column_letter(column_t_index)].width = len(t_header_value) + 2

    # Configurando a primeira linha com cabeçalhos
    setup_headers(ws, headers)

    # Processa as fases e atualiza a planilha
    row_num = 2
    ws, row_num = process_phases_nps(all_phases, headers, ws, row_num)


    print(f"Dados atualizados na aba '{sheet_name}'.")


# Função para processar os cartões da Matriz de Cursos
def process_card(card, headers):
        field_values = {header: "" for header in headers}
        field_values["Membro"] = card["title"]

        for field in card['fields']:
            if field['name'] in headers:
                value = field['value']
                try:
                    field_values[field['name']] = int(value)
                except ValueError:
                    try:
                        field_values[field['name']] = float(value.replace(',', '.'))
                    except ValueError:
                        field_values[field['name']] = clean_value(field['value'])

        return field_values

# Função para processar as fases da Matriz de Cursos
def process_phases_matriz(ws, headers, all_phases, row_num):
        # Estilo de fonte para o restante das células
        normal_font = Font(name='Arial', size=10, bold=False)
        alignment_bottom = Alignment(vertical='bottom')
        
        
        for phase in all_phases:    
            if isinstance(phase, dict) and phase.get('name') in ["In-Company", "Individual"]:  # Assegura que 'phase' é um dicionário
                for card_edge in phase['cards']['edges']:
                    card = card_edge['node']
                    field_values = process_card(card, headers)

                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=field_values[header])
                        cell.font = normal_font  # Certifique-se de que 'normal_font' está definido
                        cell.alignment = alignment_bottom  # Certifique-se de que 'alignment_bottom' está definido

                    row_num += 1


        return ws, row_num

#Função para atualizar a aba da Matriz de Cursos
def update_excel_matriz_cursos(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)


    # Lista de títulos para os cabeçalhos conforme especificado
    headers = [
        "Membro", "Qual o tipo de curso?", "Cargo atual na empresa", "Carga horária do curso", "Ajudou a desenvolver minhas soft skills", 
        "Contribuiu para o meu desenvolvimento pessoal", "Contribuiu para o meu desenvolvimento profissional",
        "Facilita meu trabalho dentro da empresa", "Fez com que meus resultados na UCJ fossem alavancados",
        "Me ajudou nas atividades que desenvolvo fora da UCJ", "Me ajudou nas atividades que desenvolvo na UCJ",
        "Pode ser utilizado no dia a dia do meu projeto", "Pode ser utilizado no meu dia a dia fora da empresa", 
        "Potencializou meu desempenho de modo geral", "Área do curso", "Nome da Instuição em que realizou o curso", "Nome do curso realizado" 
    ]

    # Aplicando os cabeçalhos e seus estilos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(vertical='bottom')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 2

    
    result = process_phases_matriz(ws, headers, all_phases, row_num)
    ws = result[0]
    row_num = result[1]


    # Ajuste das colunas conforme anteriormente
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    print(f"Dados atualizados na aba '{sheet_name}'.")


def process_card_membros(card, headers):
    # Inicializa um dicionário com todos os headers definidos para strings vazias.
    field_values = {header: "" for header in headers}

    # Atribui o título do node ao campo "Membro" especificamente.
    field_values["Membro"] = card["title"]

    # Processa os outros campos que podem estar presentes em 'fields'.
    for field in card['fields']:
        field_name = field['name']
        if field_name in headers and field_name != "Membro":
            value = field['value']
            if value is None:
                field_values[field_name] = ""
            else:
                try:
                    field_values[field_name] = int(value)
                except ValueError:
                    try:
                        field_values[field_name] = float(value.replace(',', '.'))
                    except ValueError:
                        field_values[field_name] = clean_value(value)

    return field_values

def process_phases_membros(ws, headers, all_phases, row_num):
    # Estilo de fonte para o restante das células
    normal_font = Font(name='Arial', size=10, bold=False)
    alignment_bottom = Alignment(vertical='bottom')

    for phase in all_phases:
        if isinstance(phase, dict) and phase.get('name') == "Caixa de entrada":  # Ajuste para igualdade se apenas uma fase é relevante
            for card_edge in phase['cards']['edges']:
                card = card_edge['node']
                field_values = process_card_membros(card, headers)

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=field_values[header])
                    cell.font = normal_font
                    cell.alignment = alignment_bottom

                row_num += 1

    return ws, row_num

def update_excel_painel_controle_membros(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    headers = ["Membro", "Status Trainee", "Data de Entrada (Trainee)", "Motivo de Saída (Trainee)", "Data de Entrada (Consultor)", "Data Saída", ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(vertical='bottom')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 2

     
    result = process_phases_membros(ws, headers, all_phases, row_num)
    ws = result[0]
    row_num = result[1]

    # Ajuste das colunas conforme anteriormente
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    print(f"Dados atualizados na aba '{sheet_name}'.")

# Função principal para executar o script
def main():
    # Mapeamento de pipes para suas funções de atualização específicas
    update_functions = {
        'RH - E-NPS': update_excel_e_nps,
        'RH - Matriz de Cursos': update_excel_matriz_cursos,
        'RH - Painel Controle Membros': update_excel_painel_controle_membros,
    }

    for pipe_name, (filename, sheet_name) in PIPE_TO_FILE.items():
        print(f"Iniciando a consulta dos dados do Pipefy para: {pipe_name}")
        all_phases = fetch_all_cards(PIPE_IDS[pipe_name])
        
        try:
            wb = load_workbook(filename)
            print(f"Arquivo '{filename}' carregado com sucesso.")
        except FileNotFoundError:
            wb = Workbook()
            print(f"Arquivo '{filename}' não encontrado, criando novo arquivo.")
            wb.remove(wb.active)  # Remover a aba padrão vazia

        # Chamada da função de atualização específica, agora passando todas as fases paginadas
        update_function = update_functions[pipe_name]
        update_function(wb, all_phases, sheet_name)

        wb.save(filename)
        print(f'Arquivo "{filename}" salvo com sucesso com a aba atualizada.')

if __name__ == "__main__":
    main()
