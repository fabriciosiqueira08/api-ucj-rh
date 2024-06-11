from openpyxl.styles import Font, Alignment
from ProcessCardClima import process_card_clima
from datetime import datetime

def clear_worksheet(ws):
    # Método para limpar todas as linhas e colunas da planilha
    for row in ws.iter_rows(min_row=2):  # Começa na segunda linha
        for cell in row:
            cell.value = None

def process_phases_clima(ws, headers, all_phases, row_num):

    clear_worksheet(ws)
    # Estilo de fonte para o restante das células
    normal_font = Font(name='Arial', size=10, bold=False)
    alignment_bottom = Alignment(vertical='bottom')

    for phase in all_phases:
        if isinstance(phase, dict):  
            for card_edge in phase['cards']['edges']:
                card = card_edge['node']
                field_values = process_card_clima(card, headers)

                # Acessa corretamente o campo `createdAt` dentro do node
                created_at_str = card.get('createdAt', '')
                if created_at_str:
                    try:
                        created_at = datetime.fromisoformat(created_at_str.replace("Z", "+00:00"))
                        created_at_formatted = created_at.strftime('%d/%m/%Y')
                    except ValueError:
                        created_at_formatted = created_at_str
                else:
                    created_at_formatted = ''

                # Escreve o valor `created_at` na coluna "Data:" (coluna 1)
                ws.cell(row=row_num, column=1, value=created_at_formatted)


                # Ajuste a escrita dos outros campos
                for col_num, header in enumerate(headers[1:], 2):
                    if header == "Membro":
                        cell_value = card['title']  # Obtém o título do cartão diretamente
                    else:
                        cell_value = field_values.get(header, "")
                    
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.font = normal_font
                    cell.alignment = alignment_bottom

                row_num += 1

    return ws, row_num