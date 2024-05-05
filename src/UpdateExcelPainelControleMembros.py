from ProcessPhasesMembros import process_phases_membros
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def update_excel_painel_controle_membros(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)

    headers = ["Membro", "Status Membro", "E-mail coorporativo", "Pertence a Grupo Minoritário", "Grupo Minoritário", "Data de entrada", "Seleção de lista" ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(vertical='bottom')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 2

     
    result = process_phases_membros(ws, headers, all_phases, row_num)
    ws = result[0]
    row_num = result[1]

    minor_group_col_idx = headers.index("Grupo Minoritário") + 1
    for row in range(2, row_num):
        cell = ws.cell(row=row, column=minor_group_col_idx)
        cell_value = str(cell.value) if cell.value else ""
        # Remove double quotes from the value
        cell.value = cell_value.replace('"', '')

    # Ajuste das colunas conforme anteriormente
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    print(f"Dados atualizados na aba '{sheet_name}'.")
