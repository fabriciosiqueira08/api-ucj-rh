from openpyxl.styles import Font, Alignment, Border, Side
from ProcessPhasesMatriz import process_phases_matriz
from openpyxl.utils import get_column_letter
from CleanNumeric import clean_numeric

#Função para atualizar a aba da Matriz de Cursos
def update_excel_matriz_cursos(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)


    # Lista de títulos para os cabeçalhos conforme especificado
    headers = [
        "Criação: ", "Membro", "Qual o tipo de curso?", "Cargo atual na empresa", "Carga horária do curso", "Ajudou a desenvolver minhas soft skills", 
        "Contribuiu para o meu desenvolvimento pessoal", "Contribuiu para o meu desenvolvimento profissional",
        "Facilita meu trabalho dentro da empresa", "Fez com que meus resultados na UCJ fossem alavancados",
        "Me ajudou nas atividades que desenvolvo fora da UCJ", "Me ajudou nas atividades que desenvolvo na UCJ",
        "Pode ser utilizado no dia a dia do meu projeto", "Pode ser utilizado no meu dia a dia fora da empresa", 
        "Potencializou meu desempenho de modo geral", "Área do curso", "Nome da Instuição em que realizou o curso", "Nome do curso realizado" 
    ]

    # Aplicando os cabeçalhos e seus estilos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(vertical='bottom')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 2

    
    result = process_phases_matriz(ws, headers, all_phases, row_num)
    ws = result[0]
    row_num = result[1]

# Limpeza e formatação dos dados na coluna "Carga horária do curso"
    for row in ws.iter_rows(min_row=2, max_col=5, max_row=ws.max_row, min_col=5):
        for cell in row:
            cleaned_value = clean_numeric(cell.value)
            cell.value = cleaned_value if cleaned_value is not None else "Dados inválidos"

    # Ajuste das colunas conforme anteriormente
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    print(f"Dados atualizados na aba '{sheet_name}'.")