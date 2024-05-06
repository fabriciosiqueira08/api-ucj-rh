from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ProcessPhasesClima import process_phases_clima

def update_excel_pesquisa_clima(wb, all_phases, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    headers = ["Data: ", "Membro", "Você é:", "Você deseja mais algum cargo na empresa? Se sim, qual?", "Como você avalia o quanto esteve satisfeito com o relacionamento com os membros da empresa durante este mês? *", "Comente sobre a sua satisfação com relacionamento com membros este mês.", "Como você avalia a sua satisfação em relação ao reconhecimento interno na empresa durante este mês?", "Comente sobre sua satisfação com o reconhecimento interno este mês.", "Como você avalia seu grau de satisfação em relação à geração de valor que a UCJ trouxe para você durante este mês?", "Comente sobre sua satisfação com o grau de geração de valor que a UCJ trouxe para você este mês.", "Como você avalia seu grau de satisfação no que diz respeito ao trabalho que realizou na empresa durante este mês?", "Comente sobre a sua satisfação com o trabalho realizado este mês.", "Recebo as capacitações necessárias para realizar o meu trabalho. ", "A empresa se mostra disposta a levar em consideração/ouvir minhas ideias, críticas e sugestões e a discuti-las comigo.", "Percebo que, na(s) minha(s) equipe(s), buscamos sempre nos ajudar e nos desenvolver. ", "Eu tenho autonomia o suficiente para cumprir com minhas responsabilidades da melhor forma", "Durante este mês, consegui conciliar minha vida acadêmica, profissional e pessoal.", "Sinto que minhas necessidades são entendidas e atendidas de forma ágil", "Sinto-me pertencente à UCJ e percebo que os membros me respeitam ", "Tenho todos os recursos necessários para executar bem meu trabalho", "Sinto-me livre e confortável para ser quem sou diante dos membros da empresa", "Das atividades realizadas na UCJ, qual você considera ser a mais interessante?", "Das atividades realizadas na UCJ, qual você considera ser a mais desmotivante?", "Comente sobre a atuação do seu Gerente de RH.", "As orientações que você recebe do seu diretor para a realização do seu trabalho são claras, objetivas e instrutivas?", "Comente sobre a atuação do seu superior direto."]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(vertical='bottom')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    row_num = 2

    result = process_phases_clima(ws, headers, all_phases, row_num)
    ws = result[0]
    row_num = result[1]


    # Ajuste das colunas conforme anteriormente
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width

    print(f"Dados atualizados na aba '{sheet_name}'.")